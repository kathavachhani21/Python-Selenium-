import openpyxl

# https://www.youtube.com/watch?v=nsKNPHJ9iPc
workbook = openpyxl.load_workbook('DataFile.xlsx')
print(type(workbook))

sheets = workbook.sheetnames
print(sheets)
print("active sheet name:" + workbook.active.title) # it shows the current sheet

# one method to read data
sh1 = workbook['Login'] # it's redirect to the particular sheet which we are mentioned.
data1 = sh1['B2'].value
print(data1)

# second method to read data
data2 = workbook['Advertiser']['A2'].value
print(data2)

# third method to read data
print(sh1.cell(3,1).value)
print(workbook['Advertiser'].cell(3,1).value)

# forth method to read data
print(sh1.cell(row = 2, column = 2).value)


#print(workbook.get_sheet_by_name('Login').cell(row=1, column=1).value)
row = sh1.max_row
col = sh1.max_column

print("Total rows: ",row)
print("Total cols: ", col)

for i in range(1, row+1):
    for j in range(1, col+1):
        print(sh1.cell(i,j).value)

# add the data into the excel sheet
sh1.cell(row=4, column=1, value='Katha3')
sh1.cell(row=4, column=2, value=789)
sh1.cell(row=5, column=1, value='Katha4')
sh1.cell(row=5, column=2, value=789)
workbook.save("DataFile.xlsx")

#workbook.save("name_of_file.xlsx") # it will create a new file with existing data and new data