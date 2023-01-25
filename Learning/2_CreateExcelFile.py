from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook

wb = Workbook()
wb['Sheet'].title = "Course Status"
sh1 = wb.active
sh1['A1'].value = "Name"
sh1['B1'].value = "Status"
sh1['A2'].value = "Python"
sh1['B2'].value = "On-going"
sh1['A3'].value = "API"
sh1['B3'].value = "On-Hold"
sh1['B3'].fill=PatternFill("solid",fgColor="F50707")

wb.save("Data.xlsx")